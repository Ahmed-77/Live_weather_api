import urllib3
import json
import os
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

if(os.path.exists('./weather.xlsx')):
    wb = load_workbook(filename = "weather.xlsx")
    ws = wb.active
    
	
else:
    wb = Workbook()
    dest_filename = 'weather.xlsx'
    ws.title = "S1"
    
    ws = wb.active
    
    ws.append(('City', 'Time', 'Temp','Option'))
	
http = urllib3.PoolManager()
api_key = '87b198b3d71344a785691507201207'


while(True):
	
	option1=0
	option2=0
	url ='http://api.worldweatheronline.com/premium/v1/weather.ashx?key=87b198b3d71344a785691507201207&q=New+york,ny&num_of_days=2&tp=3&format=json'
	r = http.request('GET', url)

	data=json.loads(r.data)
	city1=data['data']['request'][0]['query']

	time1= data['data']['current_condition'][0]['observation_time']
	temp1=data['data']['current_condition'][0]['temp_C']+"C /"+data['data']['current_condition'][0]['temp_F']+"F"
	list1=city1,time1,temp1,option1
	
	
	
	url ='http://api.worldweatheronline.com/premium/v1/weather.ashx?key=87b198b3d71344a785691507201207&q=London&num_of_days=2&tp=3&format=json'
	r = http.request('GET', url)
	data=json.loads(r.data)
	
	city2=data['data']['request'][0]['query']

	time2= data['data']['current_condition'][0]['observation_time']
	temp2=data['data']['current_condition'][0]['temp_C']+"C /"+data['data']['current_condition'][0]['temp_F']+"F"
	list2=city2,time2,temp2,option2
	
	flag=0
	f2=0
	for i in range(1,len(ws["A"])):
		if(ws.cell(row=i,column=1).value==city1 ):
			if(ws.cell(row=i,column=4).value==0):
				ws.cell(row=i,column=3).value = temp1
				ws.cell(row=i,column=2).value = time1
			flag=1
			wb.save("weather.xlsx")
		elif(ws.cell(row=i,column=1).value==city2  ):
			if(ws.cell(row=i,column=4).value==0):
				ws.cell(row=i,column=3).value = temp2
				ws.cell(row=i,column=2).value = time2
			f2=1
			wb.save("weather.xlsx")
		
	
	if(flag==0 and f2==0):		
		ws.append(list1)
		ws.append(list2)
		break
	elif(flag==1 and f2==1):		
		break
	elif(flag==0 and f2==1):	
		ws.append(list1)
		

	
	break
ws1=wb.create_sheet(index=1 , title = "Sheet2")
ws1.append(('no','Name of cities'))
city11=[1,city1]
city12=[2,city2]
ws1.append(city11)
ws1.append(city12)

	

wb.save("weather.xlsx")
